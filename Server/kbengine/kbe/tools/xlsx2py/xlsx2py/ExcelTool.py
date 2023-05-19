# -*- coding: gb2312 -*-
# written by kebiao, 2010/08/20
# update dependency to openpyxl by 1Pixel, 2019/12/19

import openpyxl
import os
import sys

class ExcelTool:
    """
    �򵥵ķ�װexcel���ֲ���
    ϵͳҪ�� windowsϵͳ�� ��װpython2.6�Լ�pywin32-214.win32-py2.6.exe, �Լ�ms office
    """
    def __init__(self, fileName):
        #try:
        #   self.close()
        #except:
        #   pass

        self.__workbook = None

        self.fileName = os.path.abspath(fileName)
        
    def getWorkbookEx(self, auto_create = False):
        try:
            self.__workbook = openpyxl.open(self.fileName)
            return True
        except:
            pass
            
        if auto_create:
            self.__workbook = openpyxl.Workbook()
            return True
            
        return False
        
    def getXLSX(self):
        return self.__workbook


    def close(self, saveChanges = False):
        """
        �ر�excelӦ��
        """
        
        if saveChanges:
            self.__workbook.save(self.fileName)
            
        self.__workbook.close()

    def getSheetCount(self):
        """
        ��ù��������
        """
        return len(self.__workbook.worksheets)

    def getSheetNameByIndex(self, index):
        """
        ���excel��ָ������λ���ϵı�����
        """
        return self.__workbook.worksheets[index].title

    def getSheetByIndex(self, index):
        """
        ���excel��ָ������λ���ϵı�
        """
        try:
            return self.__workbook.worksheets[index]
        except:
            return None
    def __getRowCountOnSheet(self, sheet):
        cc = sheet.max_column + 1
        canRun = True
        
        while cc > 0 and canRun:
            cc -= 1
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(i, cc).value is not None:
                    canRun = False
        return cc

    def getRowCount(self, sheetIndex):
        """
        ���һ���ж���Ԫ��
        """
        ws = self.__workbook.worksheets[sheetIndex]
        return self.__getRowCountOnSheet(ws)
        
        cc = ws.max_column + 1
        canRun = True
        
        while cc > 0 and canRun:
            cc -= 1
            for i in range(1, ws.max_row + 1):
                if ws.cell(i, cc).value is not None:
                    canRun = False
        return cc

    def getColCount(self, sheetIndex):
        """
        ���һ���ж���Ԫ��
        """
        return self.__workbook.worksheets[sheetIndex].max_row
        
    def getValue(self, sheet, row, col):
        """
        ���ĳ���������ĳ��λ���ϵ�ֵ
        """
        return sheet.cell(row, col).value

    def getText(self, sheet, row, col):
        """
        ���ĳ���������ĳ��λ���ϵ�ֵ
        """
        return str(sheet.cell(row, col).value)

    def getRowValues(self, sheet, row):
        """
        ����
        """
        cc = self.__getRowCountOnSheet(sheet)
        return [sheet.cell(row+1, i).value for i in range(1, cc + 1)]

    def getSheetRowIters(self, sheet, row):
        """
        �е�����
        """
        return sheet.Cells(1).CurrentRegion.Rows

    def getSheetColIters(self, sheet, col):
        """
        �е�����
        """
        return sheet.Cells(1).CurrentRegion.Columns

    def getColValues(self, sheet, col):
        """
        ����
        """
        rc = sheet.max_row
        return [sheet.cell(i, col+1).value for i in range(1, rc+1)]

#---------------------------------------------------------------------
#   ʹ������
#---------------------------------------------------------------------
def main():
    xbook = ExcelTool("test.xlsx")

    print("sheetCount=%i" % xbook.getSheetCount())

    for x in range(1, xbook.getSheetCount()):
       print( "      ", xbook.getSheetNameByIndex(x))

    print( "sheet1:rowCount=%i, colCount=%i" % (xbook.getRowCount(1), xbook.getColCount(1)))

    for r in range(1, xbook.getRowCount(1)+1):
        for c in range(1, xbook.getColCount(1)+1):
            val = xbook.getValue(xbook.getSheetByIndex(2), r, c)
            if val:
                print( "DATA:", val)

if __name__ == "__main__":
    main()




